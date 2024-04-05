#
# Change Log
# ----------
# 21Oct2021	David Ross
#			Tidy up, part of SW00.0139 now.
#			I am adding colour based highlight of failed values.
#			Now we take doublets of numeric value plus pass/fail into the CSV file
#			and we parse these and set colour background in each row.
#
# 25Oct2021	David Ross
#			Split __main__ from the actual working so this script can be invoked
#			directly from another script.
import os
from datetime import datetime
from openpyxl import load_workbook, styles
import csv
from xlsxwriter.workbook import Workbook


# if True: ignore if "InEar Microphone ... (L)" -> not applicable to MFP00166
MFP00166 = True
PASS_AS_ZERO = False

# change this to tuple ()?
EXPECTED_COLUMNS = [
					{"col_num" : 1,
					"value" : "NAME"},
					{"col_num" : 2,
					"value" : "MARGIN"},
					{"col_num" : 4,
					"value" : "RESULT"},
]

# will add the RHS columns in code
COLUMNS = [
    "Date",
    "Serial",
    "Result",
    "Passive Attenuation (L)",
    "Passive Attenuation (R)",
    "Loudspeaker Level (L)",
    "Loudspeaker Level (R)",
    "Loudspeaker Average Level (L) [500Hz - 2kHz]",
    "Loudspeaker Average Level (R) [500Hz - 2kHz]",
    "Loudspeaker Response (L)",
    "Loudspeaker Response (R)",
    "Loudspeaker Polarity (L)",
    "Loudspeaker Polarity (R)",
    "Loudspeaker THD (L)",
    "Loudspeaker THD (R)",
    "Loudspeaker R&B (L)",
    "Loudspeaker R&B (R)",
    "SENS Microphone Level (L)",
    "SENS Microphone Level (R)",
    "SENS Microphone Response (L)",
    "SENS Microphone Response (R)",
    "SENS Microphone THD (L)",
    "SENS Microphone THD (R)",
    "Boom Microphone Input Level",
    "Boom Microphone Input Response",
	"Boom Microphone Input THD",
]


def get_all_files(folder=""):
	path = os.path.dirname(os.path.realpath(__file__))
	path = os.path.join(path, folder)
	os.chdir(path)
	files = [os.path.join(path, file) for file in sorted(os.listdir(path), key=os.path.getmtime)
		if file.endswith(".xlsx") and 'template' not in file.lower()]
		
	return files

def sanitise(value):
	return value
	
def get_data_line(file):
	#print("get data from:".format(os.path.basename(file)), flush=True)
	wb = load_workbook(file, data_only=True)
	ws = wb["SUMMARY"]
	date = os.path.getmtime(file)
	for entry in  EXPECTED_COLUMNS:
		
		idx = entry["col_num"]
		name = entry["value"]
		if ws.cell(row=9, column=idx).value.upper() != name:
			raise Exception(f"column number '{idx}' did not match expected name '{name}'")
	
	rows = [row for row in ws.iter_rows()]
	
	data = []
	serial = ""
	date = ""
	
	# use filename to determine overall result
	fname, _ = os.path.splitext(os.path.basename(file))
	# print(fname)
	if "F" in fname[0]:
		result = "FAIL"
	else:
		result = "PASS"
		
	for i, row in enumerate(rows):

		# skip empty rows and skip the first line
		if row is not None:
			if row[3].value is not None: #what does row[9] mean? why 9?
				
				# must append in this order (left then right).
				# print(row[0].value in COLUMNS)

				if MFP00166: # and not all(x in row[0].value for x in ["InEar Microphone", "(L)"]):
					if PASS_AS_ZERO and "PASS" in row[3].value: # skip if PASS result
						data.append(0)
					elif row[0].value in COLUMNS:
						data.append(sanitise(row[1].value)) # actual numeric value
						data.append(row[3].value) # the measurement's result: 'PASS' or 'FAIL'
								
				# if PASS_AS_ZERO and "PASS" in row[10].value:  # skip if PASS result
				# 	data.append(0)
				# else:
				# 	data.append(sanitise(row[8].value))
				
			elif row[0].value is not None and "serial" in row[0].value.lower():
				# print(f"serial is: {row[1].value}\n")
				# print(row[2].value)
				serial = row[1].value
            
			elif row[0].value is not None and "date" in row[0].value.lower():
				# print(f"date is: {row[1].value}")
				date = row[1].value
			#	print(date)
				
	#print(data)	print("\n")	
	return (date, serial, result, data)
	
def processResults(path):
	files = get_all_files(path)
	csvfilename = "data.csv"

	with open(csvfilename, "w") as out:
		# setup summary spreadsheet
		for col in COLUMNS:
			# if MFP00166 and not all(x in col for x in ["InEar Microphone", "(L)"]):
			out.write(f"{col},")
			# if "(L)" in col:
			# 	col = col.replace("(L)", "(R)")
			# 	out.write(f"{col},")
				
		out.write("\n") # completed the column titles
		
		for file in files:
			print("processing: " + file, flush=True)
			if "summary" in file.lower():
				continue
			if PASS_AS_ZERO and "pass" in file.lower():
				continue
			# print(file)
			date, serial, result, data = get_data_line(file)
			#date = datetime.fromtimestamp(date0).strftime('%Y-%m-%d %H:%M:%S')
			out.write(f"{date},{serial},{result}")
			for d in data:
				out.write(f",{d}")
			out.write("\n")
			
	workbook = Workbook("Summary.xlsx")
	worksheet = workbook.add_worksheet()
	cell_format_failed = workbook.add_format({'font_color': 'Black'})
	cell_format_failed.set_bg_color('#febfb1')
	with open(csvfilename, 'rt', encoding='utf8') as f:
		reader = csv.reader(f)
		for r, row in enumerate(reader):
			#print(f"Process row {r}")
			tc = 0;				
			for c, col in enumerate(row):
				if r == 0:
					# first row - header...
					worksheet.write(r, c, col)
				else:
					# data rows...
					if c < 3:
						# date, serial and result.
						worksheet.write(r, c, col)
					else:
						# result doublets.
						if (tc % 2) == 0:
							data = float(col)
						else :
							if "FAIL" in col:
								worksheet.write_number(r, ((tc - 1)/2) + 3, data, cell_format_failed)
							else:
								worksheet.write_number(r, ((tc - 1)/2) + 3, data)
						tc = tc + 1

	cell_format = workbook.add_format({'bold': True, 'font_color': 'Black'})
	cell_format.set_bg_color('#9bf542')
	cell_format.set_border(1)
	worksheet.set_row(0, None, cell_format)
	workbook.close()

if __name__ == "__main__":
	path = input("Please provide the results path:\n> ")
	processResults(path)