###############################################################################
#
#	create data summary of MFP126, MFP166 or MFP270 results
#
#	change log
#	----------
#	25Oct2021	David Ross
#				Make this callable.
#	26Oct2021	David Ross
#				Individual results are now doublets in the data
#				{data, "PASS"/"FAIL"}
#
import os
from datetime import datetime
from openpyxl import load_workbook, styles
import csv
from xlsxwriter.workbook import Workbook

# if True: ignore if "InEar Microphone ... (L)" -> not applicable to MFP00166
MFP00166 = True
PASS_AS_ZERO = False

# change this to tuple ()?
EXPECTED_COLUMNS = [
					{"col_num" : 1,
					"value" : "NAME"},
					{"col_num" : 2,
					"value" : "MARGIN"},
					{"col_num" : 8,
					"value" : "NAME"},
					{"col_num" : 9,
					"value" : "MARGIN"},
]
# will add the RHS columns in code
COLUMNS = [
    "Date",
	"Serial",
	"Result",
	"Passive Attenuation (L)",
	"Loudspeaker Level (L)",
	"Loudspeaker Average Level (L) [500Hz - 2kHz]",
	"Loudspeaker Response (L)",
	"Loudspeaker Polarity (L)",
	"Loudspeaker THD (L)",
	"Loudspeaker R&B (L)",
	"SENS Microphone Level (L)",
	"SENS Microphone Response (L)",
	"SENS Microphone THD (L)",
	"InEar Microphone Level (L)",
	"InEar Microphone Response (L)",
	"InEar Microphone THD (L)",
	"Difference B (L)",
]


def get_all_files(folder=""):
	path = os.path.dirname(os.path.realpath(__file__))
	path = os.path.join(path, folder)
	os.chdir(path)
	files = [os.path.join(path, file) for file in sorted(os.listdir(path), key=os.path.getmtime)
		if file.endswith(".xlsx") and 'template' not in file.lower()]
		
	return files

def sanitise(value):
	return value
	
def get_data_line(file):
	wb = load_workbook(file, data_only=True)
	ws = wb["SUMMARY"]
	date = os.path.getmtime(file)
	for entry in  EXPECTED_COLUMNS:
		
		idx = entry["col_num"]
		name = entry["value"]
		#if ws[idx].value.upper() != name:
		if ws.cell(row=1, column=idx).value.upper() != name:
			raise Exception(f"column number '{idx}' did not match expected name '{name}'")
	
	rows = [row for row in ws.iter_rows()]
	
	data = []
	serial = ""
	date = ""
	
	# use filename to determine overall result
	if "PASS" in file:
		result = "PASS"
	else:
		result = "FAIL"
		
	for i, row in enumerate(rows):

		# skip empty rows and skip the first line
		if i != 0:
			if row[9].value is not None: #what does row[9] mean? why 9?
				
				# must append in this order (left then right).
				
				# if MFP00166 and not all(x in row[0].value for x in ["InEar Microphone", "(L)"]):
				if MFP00166:
					if PASS_AS_ZERO and "PASS" in row[3].value: # skip if PASS result
						data.append(0)
					else:
						data.append(sanitise(row[1].value))
						data.append(row[3].value)
								
				if PASS_AS_ZERO and "PASS" in row[10].value:  # skip if PASS result
					data.append(0)
				else:
					data.append(sanitise(row[8].value))
					data.append(row[10].value)
				
			elif row[0].value is not None and "serial" in row[0].value.lower():
				serial = row[1].value
            
			elif row[0].value is not None and "date" in row[0].value.lower():
				date = row[1].value
			#	print(date)
				
	#print(data)	print("\n")	
	return (date, serial, result, data)

	
def processResults(path):
	files = get_all_files(path)
	csvfilename = "data.csv"
	with open(csvfilename, "w") as out:
		# setup summary spreadsheet
		for col in  COLUMNS:
			# if MFP00166 and not all(x in col for x in ["InEar Microphone", "(L)"]):
			if MFP00166:
				out.write(f"{col},")
			if "(L)" in col:
				col = col.replace("(L)", "(R)")
				out.write(f"{col},")
				
		out.write("\n") # completed the column titles
		
		for file in files:
			print("processing: " + file, flush=True)
			if PASS_AS_ZERO and "pass" in file.lower():
				continue
			#print(file)
			date, serial, result, data = get_data_line(file)
			out.write(f"{date},{serial},{result}")
			for d in data:
				out.write(f",{d}")
			out.write("\n")
		out.close()
	workbook = Workbook("Summary.xlsx")
	worksheet = workbook.add_worksheet()
	cell_format_failed = workbook.add_format({'font_color': 'Black'})
	cell_format_failed.set_bg_color('#febfb1')
	with open(csvfilename, 'rt', encoding='utf8') as f:
		reader = csv.reader(f)
		for r, row in enumerate(reader):
			#print(f"Process row {r}")
			tc = 0;				
			for c, col in enumerate(row):
				if r == 0:
					# first row - header...
					worksheet.write(r, c, col)
				else:
					# data rows...
					if c < 3:
						# date, serial and result.
						worksheet.write(r, c, col)
					else:
						# result doublets.
						if (tc % 2) == 0:
							data = float(col)
						else :
							if "FAIL" in col:
								worksheet.write_number(r, ((tc - 1)/2) + 3, data, cell_format_failed)
							else:
								worksheet.write_number(r, ((tc - 1)/2) + 3, data)
						tc = tc + 1

	cell_format = workbook.add_format({'bold': True, 'font_color': 'Black'})
	cell_format.set_bg_color('#9bf542')
	cell_format.set_border(1)
	worksheet.set_row(0, None, cell_format)
	workbook.close()
	
if __name__ == "__main__":
	path = input("Please provide the results path:\n> ")
	processResults(path)